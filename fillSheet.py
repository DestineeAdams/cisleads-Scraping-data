from openpyxl import load_workbook, Workbook
import json

class fillSheet:

    def __init__(self, wbname, wsname):
        self.wb = load_workbook(wbname)
        self.ws = self.wb[wsname]
        self.nameList = []



    def fillNamesList(self):

        # get names of companies
        for cell in range(self.ws.min_row,  self.ws.max_row+1):
            currCellVal = "A"+ str(cell)
            self.nameList.append(self.ws[currCellVal].value)


    def getNamesList(self):
        self.fillNamesList()
        return self.nameList

    def updateCell(self, cell, data): 
        return

    def getSheetDim(self):
        return str(self.ws.max_row) + " x " + str(self.ws.max_column)

    def getNumOfCol(self):
        return self.ws.max_column
    
    def getNumOfRow(self):
        return self.ws.max_row
     

    


filepath = r'C:\Users\des11\OneDrive\Desktop\cisleads-Scraping-data\Perfetto Bid History Template.xlsx'
f = fillSheet(filepath, "Competitor Analysis")
# print(f.getNumOfRow())
# print(f.getSheetDim(), f.getNumOfCol(), f.getNumOfRow())

