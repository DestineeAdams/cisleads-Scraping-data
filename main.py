import os, time, re, sys
from dotenv import load_dotenv
from helium import *
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.utils import get_column_letter
import tkinter as tk


load_dotenv()  # take environment variables from .env.

info = [] # {"name":"compname", "rowNumber": "row" ,"low":0, "other":0}

headerValues = {}
html = None
compname = None
index = 10
currdictName = ""
colLow = 5 #location of the low bibs col
colLowPlusOther = 6 #location of the Low + Other

# grab excel file
filepath = r"Perfetto Bid History Template.xlsx"
sheetname = "Competitor Analysis"
    
wb = openpyxl.load_workbook(filepath)
ws = wb[sheetname]


# get sheet dimensionm
dimensions = (ws.max_row+1, ws.max_column+1)
# print(dimensions)


# get header row coordinate and row values put then in dict
for y in range (1, dimensions[1]):
    headerValues.update({ 
        ws.cell(row=1,column=y).value : ws.cell(row=1,column=y).coordinate
    })
    
# print(headerValues)
# print("--------------")

# import names of companies into dict and into list

for i in range(1, dimensions[0]):
 
    if ws.cell(row=i,column=1).value.lower() != "name":      
    
        # print(ws.cell(row=i,column=1).value)
        
        info.append({
            "name": ws.cell(row=i,column=1).value, 
            "rowNumber": i,
            "low": 0, 
            "other":0
        })


# print(info)
# currdictName = info[index]["name"] # get 

# grab table html for each name on list on site
def grabTableHtml(index):
    currdictName = info[index]["name"]
    # driver = start_firefox("https://cisleads.com/?login=1", headless=False)
    driver = start_firefox("https://cisleads.com/?login=1", headless=True)

    
    # login
    write(os.getenv("UN"), into="email:")
    write(os.getenv("PW"), into="password:")
    click(S("#buttonLogin"))
    

    # fill out form
    click(S("//html/body/section/div[2]/div[3]/div[1]/form/div[1]/a[2]"))
    write(currdictName, into=S("#Keywords"))
    
    time.sleep(1)
    click(currdictName)
    
    # click print
    click("Print")
    
    time.sleep(5) # with for page to load
    
    # save html
    html = driver.page_source
    # print(html)
    
    # end
    kill_browser()
    # print("done helium")
    
    return html

# print(html)

# put raw table html into a file
def prase():
    soup = BeautifulSoup(html, "html.parser")
    getTables = soup.find_all("table")
    Tables = [getTables[2], getTables[3]]
    
 
    
    with open("lowbids.html", "w") as f:
        data = str(Tables[0])
        f.write(data)
        
    with open("otherbids.html", "w") as f:
        data = str(Tables[1])
        f.write(data)
             
           

# prase the rawHTML.html for price of low bids
    # put sum in dict with company name
def getlowBidsSum(index):
    
    tds = None
    prices = []

    with open("lowbids.html") as fp:
        # get locbid-cell
        soup = BeautifulSoup(fp, "html.parser")
        tds = str(soup.find_all(class_="locbid-cell"))
 
        with open("low.html", "w") as f:
            f.write(tds)
    
    with open("low.html") as fq: 
        soup = BeautifulSoup(fq, "html.parser")
        strongs = list(soup.find_all("strong"))
        
        for i in range(len(strongs)):
            prices.append(strongs[i].get_text())
            
        # fill price
        for i in range(len(prices)):
            try: 
                prices[i] = int(re.sub("[$,]", "", prices[i]))
                # print(prices)

            except ValueError:
                print('A ValueError occurred getlowBidsSum')
                prices[i] = 0   
            
            except TypeError:
                print('A TypeError occurred getlowBidsSum')
                prices[i] = 0
                
     
        # print(prices)
        
    info[index]["low"] = sum(prices)
    # print(info[index])
   

# prase the html for price of Other bids
    # put sum in dict with company name
def getotherBidsSum(index):
    tds = None
    prices = []

    with open("otherbids.html") as fp:
        # get locbid-cell
        soup = BeautifulSoup(fp, "html.parser")
        tds = str(soup.find_all(class_="locbid-cell"))
 
        with open("other.html", "w") as f:
            f.write(tds)
    
    with open("other.html") as fq: 
        soup = BeautifulSoup(fq, "html.parser")
        strongs = list(soup.find_all("strong"))
        
        for i in range(len(strongs)):
            prices.append(strongs[i].get_text())
            
            
            
        for i in range(len(prices)):
            try:
                prices[i] = int(re.sub("[$,]", "", prices[i]))
                # print(prices)

            except ValueError:
                print('A ValueError occurred getotherBidsSum')
                prices[i] = 0
                
            
            except TypeError:
                print('A TypeError occurred getotherBidsSum')
                prices[i] = 0
                
        # print(prices)

    info[index]["other"] = sum(prices)
    # print(info[index])

 
#  update low excel cells
def updateLowPricesCell(col):

    for i in range(len(info)):
        letter = get_column_letter(col)

   
        # print(ws[letter + str(info[i]["rowNumber"])].value)
        ws[letter + str(info[i]["rowNumber"])] = info[i]["low"]

#  update others bids excel cells
def updateLowplusOtherbids(col):

    for i in range(len(info)):
        letter = get_column_letter(col)

   
        # print(ws[letter + str(info[i]["rowNumber"])].value)
        ws[letter + str(info[i]["rowNumber"])] = info[i]["other"] + info[i]["low"]




# # for loop to cylce thought list of name and grab html
for i in range(0, len(info)):
    
    index = i
    try:
        html = grabTableHtml(index)
        # print(html)
        
        prase()
        
        getlowBidsSum(index)
        getotherBidsSum(index)
        
        # print(f"is i: {index} result is {info[index]}")
        
        updateLowPricesCell(colLow)
        updateLowplusOtherbids(colLowPlusOther)
        wb.save(filepath)
        
        print(f" grabed {info[index]['name']} data")
        
    
    except PermissionError:
        print('PermissionError')
        print('close the excel sheet')
    
    except selenium.common.exceptions.WebDriverException:
        print('selenium.common.exceptions.WebDriverException')
        print('tring angain')
        
        html = grabTableHtml(index)
        # print(html)
        
        prase()
        
        getlowBidsSum(index)
        getotherBidsSum(index)
        
        # print(f"is i: {index} result is {info[index]}")
        
        updateLowPricesCell(colLow)
        updateLowplusOtherbids(colLowPlusOther)
        wb.save(filepath)
        
        print(f" grabed {info[index]['name']} data")
        
    
    print("COMPLETE")





#run a single time
# html = grabTableHtml(index)
# # print(html)
# prase()
# getlowBidsSum(index)
# getotherBidsSum(index)

# updateLowPricesCell(colLow)
# updateLowplusOtherbids(colLowPlusOther)

# wb.save(filepath)


