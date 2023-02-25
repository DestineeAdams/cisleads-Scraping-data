from openpyxl import load_workbook, Workbook
import json


class fillSheet:

    def __init__(self, wbname, wsname):
        self.wb = load_workbook(wbname)
        self.ws = self.wb[wsname]
        self.nameList = []


    def fillNamesList(self):
        # get names of companies
        for cell in range(self.ws.min_row,  self.ws.max_row+1):
            currCellVal = "A"+ str(cell)
            self.nameList.append(self.ws[currCellVal].value)

    def getNamesList(self):
     return self.nameList

    def getOneName(self, previousName):
        print(len(self.nameList))

        for i in range(len(self.nameList)):
            print(self.nameList[i])
            
    



f = fillSheet('Perfetto Bid History Template.xlsx', "Competitor Analysis")
f.fillNamesList()
# f.getNamesList()
f.getOneName("test")